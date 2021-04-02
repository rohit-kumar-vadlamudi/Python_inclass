import openpyxl
wb = openpyxl.load_workbook(r'D:\AIML-PGDM\Term 1\Python-1204\Data files\realestatedata.xlsx')
ws = wb.sheetnames

print("print all sheets in Excel Workbook")
for x in mysheets:
  print(x)

realestatedata = wb['stouffville']

print("print type sheet")
print(type(realestatedata))
print("sheet name") 
print(realestatedata.title)

print("Value of A1")
data1 = realestatedata['A1']
print(data1.value)

# Deposite the data in a data structure

print('Goal = Store the values of "Selling price, House Type, Description, Number of Bed rooms", into data structures')
s_price = []
houseType = []
description = []
numberBedrooms = []

for i in range(2,191):
	s_price.append(realestatedata.cell(row=i, column=4).value)
	houseType.append(realestatedata.cell(row=i, column=1).value)
	description.append(realestatedata.cell(row=i, column=2).value)
	numberBedrooms.append(realestatedata.cell(row=i, column=3).value)

s_price = tuple(s_price)
houseType = tuple(houseType)
description = tuple(description)
numberBedrooms = tuple(numberBedrooms)
print('Done!')
# Average of selling price

print("Goal = Print the average of Selling price")
sum=0
for i in s_price:
	if(i != None):
		sum+=i
	else:
		pass
print("Average Selling price is : {:.2f} dollors".format(sum/190))