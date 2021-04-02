import openpyxl
wb = openpyxl.load_workbook(r'D:\AIML-PGDM\Term 1\Python-1204\Data files\realestatedata.xlsx')
ws = wb.sheetnames

print("print all sheets in Excel Workbook")
for x in ws:
  print(x)

realestatedata = wb['stouffville']

print("print type sheet")
print(type(realestatedata))
print("sheet name") 
print(realestatedata.title)

print("Value of A1")
data1 = realestatedata['A1']
print(data1.value)

for i in range(1, 191):
	print(realestatedata.cell(row=i, column=10).value)