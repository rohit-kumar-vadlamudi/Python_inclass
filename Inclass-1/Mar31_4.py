import openpyxl
from array import *

wb = openpyxl.load_workbook(r'D:\AIML-PGDM\Term 1\Python-1204\Data files\realestatedata.xlsx')
ws = wb.sheetnames

print("print all sheets in Excel Workbook")
for x in ws:
  print(x)

realestatedata = wb['stouffville']

print("print type sheet")
print(type(realestatedata))
print("sheet name") 
print(realestatedata.title)

WhitbyREData = ["start"]

for i in range(1, 191):
	price = realestatedata.cell(row=i, column=4).value; 
	houseType = realestatedata.cell(row=i, column=1).value; 
	description = realestatedata.cell(row=i, column=2).value; 
	numberBedrooms = realestatedata.cell(row=i, column=3).value;

	solddetails = (houseType, description,numberBedrooms, price)
	WhitbyREData.append(solddetails)


for x in WhitbyREData:
	# print(x[0], x[1], x[2], x[3])
	if x[0] == ("Detached"):	
		print(x[3]) 