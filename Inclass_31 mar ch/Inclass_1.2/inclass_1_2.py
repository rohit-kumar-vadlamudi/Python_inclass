import openpyxl
wb = openpyxl.load_workbook('realestatedata.xlsx')
mysheets = wb.sheetnames

print(print all sheets in Excel Workbook)
for x in mysheets
  print(x)

realestatedata = wb['stouffville']

print(print type sheet)
print(type(realestatedata))
print(sheet name) 
print(realestatedata.title)

print(Value of A1)
data1 = realestatedata['A1']
print(data1.value)

for i in range(1, 190)
	print(realestatedata.cell(row=i, column=10).value)